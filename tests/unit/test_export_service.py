"""
Tests for export service facade and sub-services.

Tests DataExportService (CSV, JSON, Excel, ZIP archive),
ProtocolExportService (text, CSV, PDF), and
FigureExportService (PNG, SVG, PDF figure export).
"""
import csv
import json
import zipfile
from datetime import datetime, timezone
from io import BytesIO, StringIO
from unittest.mock import MagicMock, patch

import pytest

from app.services.data_export_service import DataExportService
from app.services.figure_export_service import FigureExportService
from app.services.protocol_export_service import ProtocolExportService
from app.services.export_service import ExportService


# ============================================================
# TestDataExportCSV
# ============================================================

class TestDataExportCSV:
    """Tests for DataExportService.export_data_csv."""

    def test_export_csv_basic(self):
        """Basic CSV export from list of dicts produces correct output."""
        data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25},
        ]
        result = DataExportService.export_data_csv(data)
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert len(rows) == 2
        assert rows[0]["name"] == "Alice"
        assert rows[0]["age"] == "30"
        assert rows[1]["name"] == "Bob"

    def test_export_csv_with_columns(self):
        """Custom column order is respected and extra keys are ignored."""
        data = [
            {"name": "Alice", "age": 30, "city": "NYC"},
            {"name": "Bob", "age": 25, "city": "LA"},
        ]
        result = DataExportService.export_data_csv(data, columns=["age", "name"])
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        # Only the specified columns should be in the output
        assert list(rows[0].keys()) == ["age", "name"]
        assert rows[0]["age"] == "30"
        assert rows[0]["name"] == "Alice"

    def test_export_csv_empty_data(self):
        """Empty data list returns empty string."""
        result = DataExportService.export_data_csv([])
        assert result == ""

    def test_export_csv_special_characters(self):
        """Commas, quotes, and newlines in values are handled properly."""
        data = [
            {"desc": 'has "quotes"', "value": "has, comma"},
            {"desc": "has\nnewline", "value": "normal"},
        ]
        result = DataExportService.export_data_csv(data)
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert rows[0]["desc"] == 'has "quotes"'
        assert rows[0]["value"] == "has, comma"
        assert rows[1]["desc"] == "has\nnewline"

    def test_export_csv_missing_keys(self):
        """Rows with missing keys produce empty values; extra keys are ignored with custom columns."""
        data = [
            {"a": 1, "b": 2, "c": 3},
            {"a": 4, "b": 5},  # missing "c"
        ]
        # Use columns to specify desired output; extrasaction='ignore' handles extra keys
        result = DataExportService.export_data_csv(data, columns=["a", "b", "c"])
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert rows[0]["c"] == "3"
        assert rows[1]["c"] == ""  # missing key becomes empty

    def test_export_csv_single_row(self):
        """Single row data produces header + one data line."""
        data = [{"x": 10}]
        result = DataExportService.export_data_csv(data)
        lines = result.strip().split("\n")
        assert len(lines) == 2
        assert "x" in lines[0]
        assert "10" in lines[1]

    def test_export_csv_numeric_values(self):
        """Numeric values including floats are serialized correctly."""
        data = [{"int_val": 42, "float_val": 3.14159, "neg": -1.5}]
        result = DataExportService.export_data_csv(data)
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert rows[0]["int_val"] == "42"
        assert rows[0]["float_val"] == "3.14159"
        assert rows[0]["neg"] == "-1.5"

    def test_export_csv_none_values(self):
        """None values are serialized as empty strings by csv.DictWriter."""
        data = [{"a": None, "b": "ok"}]
        result = DataExportService.export_data_csv(data)
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert rows[0]["a"] == ""
        assert rows[0]["b"] == "ok"


# ============================================================
# TestDataExportJSON
# ============================================================

class TestDataExportJSON:
    """Tests for DataExportService.export_data_json."""

    def test_export_json_basic(self):
        """Basic JSON serialization of dict and list."""
        data = {"key": "value", "count": 42}
        result = DataExportService.export_data_json(data)
        parsed = json.loads(result)
        assert parsed["key"] == "value"
        assert parsed["count"] == 42

    def test_export_json_datetime(self):
        """Datetime objects are serialized as ISO format strings."""
        dt = datetime(2026, 2, 22, 10, 30, 0)
        data = {"timestamp": dt}
        result = DataExportService.export_data_json(data)
        parsed = json.loads(result)
        assert parsed["timestamp"] == "2026-02-22T10:30:00"

    def test_export_json_indent(self):
        """Custom indentation is applied."""
        data = {"a": 1}
        result_4 = DataExportService.export_data_json(data, indent=4)
        result_0 = DataExportService.export_data_json(data, indent=0)
        # indent=4 should have 4-space indentation
        assert "    " in result_4
        # indent=0 should have newlines but no leading spaces for keys
        assert "    " not in result_0

    def test_export_json_nested(self):
        """Nested data structures serialize correctly."""
        data = {
            "metadata": {"version": "1.0", "tags": ["a", "b"]},
            "values": [1, 2, 3],
        }
        result = DataExportService.export_data_json(data)
        parsed = json.loads(result)
        assert parsed["metadata"]["tags"] == ["a", "b"]
        assert parsed["values"] == [1, 2, 3]

    def test_export_json_non_serializable(self):
        """Non-serializable types raise TypeError."""
        data = {"obj": object()}
        with pytest.raises(TypeError):
            DataExportService.export_data_json(data)


# ============================================================
# TestDataExportFilename
# ============================================================

class TestDataExportFilename:
    """Tests for DataExportService.generate_filename."""

    def test_generate_filename_basic(self):
        """Base name + extension generates filename with timestamp."""
        result = DataExportService.generate_filename("results", "csv")
        assert result.startswith("results_")
        assert result.endswith(".csv")

    def test_generate_filename_with_project(self):
        """Project name is included and sanitized."""
        result = DataExportService.generate_filename(
            "export", "json", project_name="My Project"
        )
        # Spaces become underscores
        assert result.startswith("My_Project_export_")
        assert result.endswith(".json")

    def test_generate_filename_no_timestamp(self):
        """timestamp=False produces filename without date suffix."""
        result = DataExportService.generate_filename(
            "data", "xlsx", timestamp=False
        )
        assert result == "data.xlsx"

    def test_generate_filename_special_chars(self):
        """Special characters in project name are sanitized to underscores."""
        result = DataExportService.generate_filename(
            "report", "pdf",
            project_name="Test/Project (v2.0)!",
            timestamp=False,
        )
        # Only alphanumeric, hyphens, underscores should remain
        assert "/" not in result
        assert "(" not in result
        assert "!" not in result
        assert result.endswith(".pdf")
        # The project name part should have sanitized chars
        assert result.startswith("Test_Project__v2_0__")


# ============================================================
# TestDataExportExcel
# ============================================================

class TestDataExportExcel:
    """Tests for DataExportService.export_excel_multisheet and export_analysis_excel."""

    def test_export_excel_multisheet(self):
        """Creates Excel workbook with multiple named sheets."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        sheets = {
            "Sheet A": [{"x": 1, "y": 2}, {"x": 3, "y": 4}],
            "Sheet B": [{"name": "test"}],
        }
        result = DataExportService.export_excel_multisheet(sheets)
        assert isinstance(result, bytes)

        # Read back with openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Sheet A" in wb.sheetnames
        assert "Sheet B" in wb.sheetnames

        ws_a = wb["Sheet A"]
        # Header row + 2 data rows
        assert ws_a.cell(1, 1).value == "x"
        assert ws_a.cell(1, 2).value == "y"
        assert ws_a.cell(2, 1).value == 1
        assert ws_a.cell(2, 2).value == 2

    def test_export_excel_with_metadata(self):
        """Metadata is written to a 'Metadata' sheet when provided."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        sheets = {"Data": [{"col": "val"}]}
        metadata = {"Author": "TestUser", "Version": "1.0"}
        result = DataExportService.export_excel_multisheet(sheets, metadata=metadata)

        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Metadata" in wb.sheetnames
        assert "Data" in wb.sheetnames

        # Metadata sheet should contain our keys
        ws_meta = wb["Metadata"]
        cell_values = []
        for row in ws_meta.iter_rows(values_only=True):
            cell_values.append(row)
        # Find the Author key
        found_author = any("Author" in str(row) for row in cell_values)
        assert found_author

    def test_export_analysis_excel(self):
        """Comprehensive analysis export creates expected sheets."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        raw_data = [{"well": "A1", "time": 0, "fluor": 100}]
        fitted = [{"well": "A1", "k_obs": 0.05, "r_squared": 0.99}]
        fold_changes = [
            {"construct": "Mutant_A", "mean_fold": 2.1},
            {"construct": "Mutant_B", "mean_fold": 1.5},
        ]

        result = DataExportService.export_analysis_excel(
            raw_data=raw_data,
            fitted_params=fitted,
            fold_changes=fold_changes,
        )
        assert isinstance(result, bytes)

        wb = openpyxl.load_workbook(BytesIO(result))
        # Should have Metadata + Raw Data + Fitted Parameters + Fold Changes
        assert "Metadata" in wb.sheetnames
        assert "Raw Data" in wb.sheetnames
        assert "Fitted Parameters" in wb.sheetnames
        assert "Fold Changes" in wb.sheetnames

    def test_export_excel_empty_sheets(self):
        """Empty data lists produce sheets without data rows."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        sheets = {"Empty": []}
        result = DataExportService.export_excel_multisheet(sheets)
        assert isinstance(result, bytes)

        wb = openpyxl.load_workbook(BytesIO(result))
        # The sheet exists but has no data rows
        ws = wb["Empty"]
        # Should have no rows of data (max_row might be None or 0 for empty sheet)
        assert ws.max_row is None or ws.max_row <= 1


# ============================================================
# TestDataExportJsonArchive
# ============================================================

class TestDataExportJsonArchive:
    """Tests for DataExportService.export_json_archive."""

    def test_export_json_archive_basic(self):
        """Creates ZIP containing separate JSON files for each data key."""
        data = {
            "raw_data": [{"well": "A1", "value": 100}],
            "params": {"k_obs": 0.05},
        }
        result = DataExportService.export_json_archive(data)
        assert isinstance(result, bytes)

        with zipfile.ZipFile(BytesIO(result), "r") as zf:
            names = zf.namelist()
            assert "metadata.json" in names
            assert "raw_data.json" in names
            assert "params.json" in names

            # Verify content
            raw = json.loads(zf.read("raw_data.json"))
            assert raw == [{"well": "A1", "value": 100}]

    def test_export_json_archive_with_metadata(self):
        """Metadata JSON includes software info and contents list."""
        data = {"results": [1, 2, 3]}
        result = DataExportService.export_json_archive(data, include_metadata=True)

        with zipfile.ZipFile(BytesIO(result), "r") as zf:
            meta = json.loads(zf.read("metadata.json"))
            assert meta["software"] == "IVT Kinetics Analyzer"
            assert meta["version"] == "1.0.0"
            assert "results" in meta["contents"]

    def test_export_json_archive_without_metadata(self):
        """When include_metadata=False, no metadata.json is created."""
        data = {"values": [1]}
        result = DataExportService.export_json_archive(data, include_metadata=False)

        with zipfile.ZipFile(BytesIO(result), "r") as zf:
            assert "metadata.json" not in zf.namelist()
            assert "values.json" in zf.namelist()

    def test_export_json_archive_datetime(self):
        """Datetime objects in data are serialized as ISO strings in the archive."""
        dt = datetime(2026, 1, 15, 12, 0, 0)
        data = {"events": [{"time": dt, "type": "start"}]}
        result = DataExportService.export_json_archive(data)

        with zipfile.ZipFile(BytesIO(result), "r") as zf:
            events = json.loads(zf.read("events.json"))
            assert events[0]["time"] == "2026-01-15T12:00:00"


# ============================================================
# TestProtocolExport
# ============================================================

class TestProtocolExport:
    """Tests for ProtocolExportService."""

    def test_export_protocol_text(self):
        """export_protocol_text delegates to format_protocol_text."""
        mock_protocol = MagicMock()
        with patch(
            "app.services.protocol_export_service.format_protocol_text",
            return_value="Protocol Text",
        ) as mock_fmt:
            result = ProtocolExportService.export_protocol_text(mock_protocol)
            mock_fmt.assert_called_once_with(mock_protocol)
            assert result == "Protocol Text"

    def test_export_protocol_csv(self):
        """export_protocol_csv delegates to format_protocol_csv."""
        mock_protocol = MagicMock()
        with patch(
            "app.services.protocol_export_service.format_protocol_csv",
            return_value="col1,col2\nval1,val2\n",
        ) as mock_fmt:
            result = ProtocolExportService.export_protocol_csv(mock_protocol)
            mock_fmt.assert_called_once_with(mock_protocol)
            assert "col1" in result

    def test_export_protocol_pdf_fallback(self):
        """When reportlab is not available, falls back to simple text-based PDF."""
        mock_protocol = MagicMock()

        with patch(
            "app.services.protocol_export_service.format_protocol_text",
            return_value="Simple protocol text",
        ):
            with patch("importlib.util.find_spec", return_value=None):
                result = ProtocolExportService.export_protocol_pdf(mock_protocol)
                assert isinstance(result, bytes)
                assert result == b"Simple protocol text"

    def test_export_protocol_pdf_simple(self):
        """_generate_pdf_simple encodes protocol text as UTF-8 bytes."""
        mock_protocol = MagicMock()
        with patch(
            "app.services.protocol_export_service.format_protocol_text",
            return_value="Test protocol content",
        ):
            result = ProtocolExportService._generate_pdf_simple(mock_protocol)
            assert isinstance(result, bytes)
            assert result.decode("utf-8") == "Test protocol content"


# ============================================================
# TestFigureExport
# ============================================================

class TestFigureExport:
    """Tests for FigureExportService."""

    def test_export_figure_png(self):
        """export_figure_png calls fig.to_image with correct PNG params."""
        mock_fig = MagicMock()
        mock_fig.to_image.return_value = b"\x89PNG\r\n"

        result = FigureExportService.export_figure_png(mock_fig, width=800, height=600, scale=2.0)
        mock_fig.to_image.assert_called_once_with(
            format="png", width=800, height=600, scale=2.0
        )
        assert result == b"\x89PNG\r\n"

    def test_export_figure_svg(self):
        """export_figure_svg calls fig.to_image for SVG and decodes to string."""
        mock_fig = MagicMock()
        mock_fig.to_image.return_value = b'<svg xmlns="http://www.w3.org/2000/svg"></svg>'

        result = FigureExportService.export_figure_svg(mock_fig, width=1000, height=700)
        mock_fig.to_image.assert_called_once_with(
            format="svg", width=1000, height=700
        )
        assert isinstance(result, str)
        assert "<svg" in result

    def test_export_figure_pdf(self):
        """export_figure_pdf calls fig.to_image for PDF format."""
        mock_fig = MagicMock()
        mock_fig.to_image.return_value = b"%PDF-1.4"

        result = FigureExportService.export_figure_pdf(mock_fig, width=1200, height=800)
        mock_fig.to_image.assert_called_once_with(
            format="pdf", width=1200, height=800
        )
        assert result == b"%PDF-1.4"

    def test_export_figure_png_raises_on_failure(self):
        """RuntimeError is raised when fig.to_image fails."""
        mock_fig = MagicMock()
        mock_fig.to_image.side_effect = ValueError("kaleido not found")

        with pytest.raises(RuntimeError, match="PNG export failed"):
            FigureExportService.export_figure_png(mock_fig)

    def test_export_figure_svg_raises_on_failure(self):
        """RuntimeError is raised when SVG export fails."""
        mock_fig = MagicMock()
        mock_fig.to_image.side_effect = ValueError("export error")

        with pytest.raises(RuntimeError, match="SVG export failed"):
            FigureExportService.export_figure_svg(mock_fig)

    def test_export_figure_pdf_raises_on_failure(self):
        """RuntimeError is raised when PDF export fails."""
        mock_fig = MagicMock()
        mock_fig.to_image.side_effect = ValueError("export error")

        with pytest.raises(RuntimeError, match="PDF export failed"):
            FigureExportService.export_figure_pdf(mock_fig)

    def test_get_figure_dimensions(self):
        """get_figure_dimensions returns width and height from layout."""
        mock_fig = MagicMock()
        mock_fig.layout.width = 1024
        mock_fig.layout.height = 768

        dims = FigureExportService.get_figure_dimensions(mock_fig)
        assert dims == {"width": 1024, "height": 768}

    def test_get_figure_dimensions_defaults(self):
        """Defaults to 700x450 when layout dimensions are None."""
        mock_fig = MagicMock()
        mock_fig.layout.width = None
        mock_fig.layout.height = None

        dims = FigureExportService.get_figure_dimensions(mock_fig)
        assert dims == {"width": 700, "height": 450}

    def test_export_figure_png_default_params(self):
        """Default parameters for PNG export are correct."""
        mock_fig = MagicMock()
        mock_fig.to_image.return_value = b"png_data"

        FigureExportService.export_figure_png(mock_fig)
        mock_fig.to_image.assert_called_once_with(
            format="png", width=1200, height=800, scale=2.5
        )


# ============================================================
# TestExportServiceFacade
# ============================================================

class TestExportServiceFacade:
    """Tests for the ExportService facade that delegates to sub-services."""

    def test_facade_delegates_csv(self):
        """ExportService.export_data_csv delegates to DataExportService."""
        data = [{"a": 1}]
        facade_result = ExportService.export_data_csv(data)
        direct_result = DataExportService.export_data_csv(data)
        assert facade_result == direct_result

    def test_facade_delegates_json(self):
        """ExportService.export_data_json delegates to DataExportService."""
        data = {"key": "value"}
        facade_result = ExportService.export_data_json(data)
        direct_result = DataExportService.export_data_json(data)
        assert facade_result == direct_result

    def test_facade_has_all_methods(self):
        """ExportService exposes all expected methods from sub-services."""
        expected_methods = [
            # Protocol exports
            "export_protocol_text",
            "export_protocol_csv",
            "export_protocol_pdf",
            # Figure exports
            "export_figure_png",
            "export_figure_svg",
            "export_figure_pdf",
            "export_figures_combined_pdf",
            "export_figures_separate_pdfs",
            "get_figure_dimensions",
            # Data exports
            "export_data_csv",
            "export_data_json",
            "generate_filename",
            "export_excel_multisheet",
            "export_analysis_excel",
            "export_json_archive",
            "export_results_summary_csv",
            "export_plate_data_csv",
            "get_raw_data_for_export",
            "get_results_for_export",
            "get_mcmc_traces_for_export",
        ]
        for method_name in expected_methods:
            assert hasattr(ExportService, method_name), (
                f"ExportService missing method: {method_name}"
            )
            assert callable(getattr(ExportService, method_name))

    def test_facade_delegates_filename(self):
        """ExportService.generate_filename delegates to DataExportService."""
        facade_result = ExportService.generate_filename(
            "test", "csv", timestamp=False
        )
        direct_result = DataExportService.generate_filename(
            "test", "csv", timestamp=False
        )
        assert facade_result == direct_result

    def test_facade_delegates_figure(self):
        """ExportService.export_figure_png delegates to FigureExportService."""
        mock_fig = MagicMock()
        mock_fig.to_image.return_value = b"png"

        facade_result = ExportService.export_figure_png(mock_fig)
        assert facade_result == b"png"


# ============================================================
# TestDataExportResultsSummaryCSV
# ============================================================

class TestDataExportResultsSummaryCSV:
    """Tests for DataExportService.export_results_summary_csv."""

    def test_results_summary_csv_basic(self):
        """Basic fold change summary CSV export."""
        fold_changes = [
            {
                "construct": "Mutant_A",
                "family": "Tbox1",
                "reference": "WT",
                "mean_log2": 1.5,
                "mean_fold": 2.83,
                "n_replicates": 6,
            },
        ]
        result = DataExportService.export_results_summary_csv(fold_changes)
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["construct"] == "Mutant_A"

    def test_results_summary_csv_empty(self):
        """Empty fold changes returns empty string."""
        result = DataExportService.export_results_summary_csv([])
        assert result == ""

    def test_results_summary_csv_without_ci(self):
        """include_ci=False excludes confidence interval columns."""
        fold_changes = [
            {
                "construct": "Mutant_A",
                "family": "Tbox1",
                "reference": "WT",
                "mean_log2": 1.5,
                "mean_fold": 2.83,
                "ci_lower_log2": 1.0,
                "ci_upper_log2": 2.0,
                "ci_lower_fold": 2.0,
                "ci_upper_fold": 4.0,
                "n_replicates": 6,
            },
        ]
        result = DataExportService.export_results_summary_csv(
            fold_changes, include_ci=False
        )
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert "ci_lower_log2" not in rows[0]


# ============================================================
# TestDataExportPlateDataCSV
# ============================================================

class TestDataExportPlateDataCSV:
    """Tests for DataExportService.export_plate_data_csv."""

    def test_plate_data_long_format(self):
        """Long format produces one row per well."""
        plate_data = {
            "wells": {
                "A1": {"construct": "WT", "type": "test"},
                "A2": {"construct": "Mutant", "type": "test"},
            }
        }
        result = DataExportService.export_plate_data_csv(plate_data, format_type="long")
        reader = csv.DictReader(StringIO(result))
        rows = list(reader)
        assert len(rows) == 2
        well_ids = {row["well_id"] for row in rows}
        assert "A1" in well_ids
        assert "A2" in well_ids

    def test_plate_data_wide_format(self):
        """Wide format produces time series with wells as columns."""
        plate_data = {
            "time_points": [0, 60, 120],
            "wells": {
                "A1": {"measurements": [100, 200, 300]},
                "A2": {"measurements": [150, 250, 350]},
            },
        }
        result = DataExportService.export_plate_data_csv(plate_data, format_type="wide")
        lines = result.strip().split("\n")
        # Header + 3 time points
        assert len(lines) == 4
        header = lines[0]
        assert "time" in header
        assert "A1" in header
        assert "A2" in header

    def test_plate_data_empty_wells(self):
        """Empty wells dict produces empty output."""
        plate_data = {"wells": {}}
        result = DataExportService.export_plate_data_csv(plate_data, format_type="long")
        assert result.strip() == ""
