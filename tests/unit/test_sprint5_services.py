"""
Tests for Sprint 5 services.

Tests:
- Methods text generation
- Publication package generation
- Package validation
- Data export (Excel, JSON)
"""
import pytest
import json
import math
from datetime import datetime
from pathlib import Path
from io import BytesIO
import tempfile
import zipfile

from app.services.methods_text_service import (
    MethodsTextService,
    MethodsTextConfig,
)
from app.services.publication_package_service import (
    PublicationPackageService,
    PublicationPackageConfig,
    FileHash,
)
from app.services.validation_service import (
    ValidationService,
    ValidationResult,
    ValidationCertificate,
    DiffReport,
)
from app.services.export_service import ExportService


# ========== Methods Text Service Tests ==========

class TestMethodsTextConfig:
    """Tests for MethodsTextConfig."""

    def test_default_config(self):
        """Test default configuration values."""
        config = MethodsTextConfig()

        assert config.n_samples == 4000
        assert config.n_chains == 4
        assert config.n_warmup == 1000
        assert config.target_accept == 0.8
        assert config.model_type == "hierarchical_bayesian"
        assert config.ci_level == 0.95
        assert config.wt_constructs == []

    def test_custom_config(self):
        """Test custom configuration values."""
        config = MethodsTextConfig(
            n_samples=2000,
            n_chains=2,
            n_constructs=50,
            n_plates=10,
        )

        assert config.n_samples == 2000
        assert config.n_chains == 2
        assert config.n_constructs == 50
        assert config.n_plates == 10


class TestMethodsTextService:
    """Tests for MethodsTextService."""

    def test_generate_data_collection_section(self):
        """Test data collection section generation."""
        config = MethodsTextConfig(
            n_constructs=25,
            n_plates=8,
            n_sessions=3,
            n_wells=200,
            has_unregulated=True,
            wt_constructs=["WT1", "WT2"],
        )

        text = MethodsTextService.generate_data_collection_section(config)

        assert "25 constructs" in text
        assert "8 plates" in text
        assert "3 independent sessions" in text
        assert "200 wells" in text
        assert "unregulated reference" in text
        assert "WT1" in text
        assert "WT2" in text

    def test_generate_curve_fitting_section(self):
        """Test curve fitting section generation."""
        config = MethodsTextConfig(
            fitting_method="nonlinear_least_squares",
            fitting_algorithm="Levenberg-Marquardt",
            r_squared_threshold=0.95,
        )

        text = MethodsTextService.generate_curve_fitting_section(config)

        assert "first-order kinetic model" in text
        assert "Levenberg-Marquardt" in text
        assert "R² < 0.95" in text
        assert "log₂-transformed" in text

    def test_generate_statistical_analysis_section(self):
        """Test statistical analysis section generation."""
        config = MethodsTextConfig(
            n_chains=4,
            n_warmup=1000,
            n_samples=4000,
            target_accept=0.8,
            ci_level=0.95,
        )

        text = MethodsTextService.generate_statistical_analysis_section(config)

        assert "4 chains" in text
        assert "1000 warmup" in text
        assert "4000 sampling" in text
        assert "95%" in text
        assert "VIF" in text

    def test_generate_software_section(self):
        """Test software section generation."""
        config = MethodsTextConfig(
            software_name="IVT Kinetics Analyzer",
            software_version="1.0.0",
        )

        text = MethodsTextService.generate_software_section(config)

        assert "IVT Kinetics Analyzer" in text
        assert "version 1.0.0" in text

    def test_generate_full_methods(self):
        """Test full methods generation."""
        config = MethodsTextConfig(
            n_constructs=25,
            n_plates=8,
        )

        text = MethodsTextService.generate_full_methods(config)

        assert "**Data Collection**" in text
        assert "**Curve Fitting**" in text
        assert "**Statistical Analysis**" in text
        assert "**Software**" in text

    def test_generate_full_methods_without_software(self):
        """Test full methods without software section."""
        config = MethodsTextConfig()

        text = MethodsTextService.generate_full_methods(config, include_software=False)

        assert "**Data Collection**" in text
        assert "**Software**" not in text

    def test_generate_methods_from_analysis(self):
        """Test methods generation from analysis results."""
        analysis_result = {
            "mcmc_config": {
                "n_samples": 2000,
                "n_chains": 2,
            },
            "data_summary": {
                "n_constructs": 10,
                "n_plates": 4,
            },
            "convergence": {
                "r_hat_threshold": 1.05,
            },
        }

        text = MethodsTextService.generate_methods_from_analysis(analysis_result)

        assert "2000 sampling" in text
        assert "2 chains" in text
        assert "10 constructs" in text

    def test_generate_latex_methods(self):
        """Test LaTeX formatting."""
        config = MethodsTextConfig()

        latex = MethodsTextService.generate_latex_methods(config)

        assert "$k_{obs}$" in latex or "k_obs" in latex
        assert "$R^2$" in latex or "R²" in latex

    def test_generate_citation(self):
        """Test citation generation."""
        citation = MethodsTextService.generate_citation()

        assert "IVT Kinetics Analyzer" in citation
        assert datetime.now().strftime("%Y") in citation


# ========== Publication Package Service Tests ==========

class TestPublicationPackageConfig:
    """Tests for PublicationPackageConfig."""

    def test_default_config(self):
        """Test default configuration values."""
        config = PublicationPackageConfig()

        assert config.title == ""
        assert config.authors == []
        assert config.license == "CC-BY-4.0"
        assert config.include_raw_data is True
        assert config.include_mcmc_traces is True
        assert config.figure_format == "png"

    def test_custom_config(self):
        """Test custom configuration values."""
        config = PublicationPackageConfig(
            title="Test Analysis",
            authors=["Author 1", "Author 2"],
            keywords=["kinetics", "IVT"],
        )

        assert config.title == "Test Analysis"
        assert len(config.authors) == 2
        assert "kinetics" in config.keywords


class TestPublicationPackageService:
    """Tests for PublicationPackageService."""

    def test_compute_file_hash(self):
        """Test SHA-256 hash computation."""
        data = b"test data"
        hash1 = PublicationPackageService.compute_file_hash(data)
        hash2 = PublicationPackageService.compute_file_hash(data)

        # Same data should produce same hash
        assert hash1 == hash2
        # Hash should be 64 characters (256 bits = 32 bytes = 64 hex chars)
        assert len(hash1) == 64

        # Different data should produce different hash
        hash3 = PublicationPackageService.compute_file_hash(b"other data")
        assert hash1 != hash3

    def test_create_package_structure(self):
        """Test directory structure creation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            base_path = Path(tmpdir) / "test_package"
            paths = PublicationPackageService.create_package_structure(base_path)

            assert "raw_data" in paths
            assert "processed_data" in paths
            assert "figures" in paths
            assert "mcmc_traces" in paths
            assert "metadata" in paths
            assert "audit" in paths

            # Verify directories exist
            for path in paths.values():
                assert path.exists()
                assert path.is_dir()

    def test_generate_manifest(self):
        """Test manifest generation."""
        files = [
            FileHash("file1.json", "abc123", 100, "application/json"),
            FileHash("file2.csv", "def456", 200, "text/csv"),
        ]
        config = PublicationPackageConfig(
            title="Test Package",
            authors=["Test Author"],
        )

        manifest = PublicationPackageService.generate_manifest(
            files,
            config,
            analysis_config={"n_samples": 4000},
            software_info={"name": "Test", "version": "1.0"},
        )

        assert manifest["manifest_version"] == "1.0"
        assert manifest["package"]["title"] == "Test Package"
        assert len(manifest["files"]) == 2
        assert manifest["checksums"]["total_files"] == 2
        assert manifest["checksums"]["total_size_bytes"] == 300

    def test_generate_datacite_metadata(self):
        """Test DataCite metadata generation."""
        config = PublicationPackageConfig(
            title="Test Dataset",
            authors=["Author One", "Author Two"],
            keywords=["kinetics", "enzymes"],
            license="CC-BY-4.0",
        )

        metadata = PublicationPackageService.generate_datacite_metadata(
            config,
            {"n_constructs": 10, "n_plates": 5},
        )

        assert metadata["titles"][0]["title"] == "Test Dataset"
        assert len(metadata["creators"]) == 2
        assert len(metadata["subjects"]) == 2
        assert metadata["publicationYear"] == datetime.now().year

    def test_export_raw_data(self):
        """Test raw data export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir)
            raw_data = {
                "plates": {
                    "plate1": {"wells": {"A1": [1, 2, 3]}},
                    "plate2": {"wells": {"A1": [4, 5, 6]}},
                },
            }

            file_hashes = PublicationPackageService.export_raw_data(
                raw_data, output_path
            )

            assert len(file_hashes) == 2
            assert all(h.content_type == "application/json" for h in file_hashes)
            assert (output_path / "plate_plate1_raw.json").exists()
            assert (output_path / "plate_plate2_raw.json").exists()

    def test_export_processed_results(self):
        """Test processed results export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir)
            results = {
                "fitted_params": [
                    {"well_id": "A1", "k_obs": 0.1, "f_max": 100},
                    {"well_id": "A2", "k_obs": 0.2, "f_max": 150},
                ],
                "fold_changes": [
                    {"construct": "MUT1", "mean": 2.0, "ci_lower": 1.5, "ci_upper": 2.5},
                ],
                "posterior_summary": {"mean": [1.0, 2.0]},
            }

            file_hashes = PublicationPackageService.export_processed_results(
                results, output_path
            )

            assert len(file_hashes) >= 2
            assert (output_path / "fitted_parameters.csv").exists()
            assert (output_path / "fold_changes.csv").exists()

    def test_generate_audit_log(self):
        """Test audit log generation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir)
            events = [
                {"timestamp": "2024-01-01T10:00:00", "action": "Upload", "user": "Test"},
                {"timestamp": "2024-01-01T11:00:00", "action": "Analysis", "user": "Test"},
            ]

            file_hashes = PublicationPackageService.generate_audit_log(
                events, output_path
            )

            assert len(file_hashes) == 2
            assert (output_path / "audit_log.json").exists()
            assert (output_path / "audit_log.md").exists()

            # Verify JSON content
            with open(output_path / "audit_log.json") as f:
                audit_json = json.load(f)
                assert len(audit_json["events"]) == 2

    def test_create_publication_package(self):
        """Test complete package creation."""
        config = PublicationPackageConfig(
            title="Test Analysis",
            authors=["Test Author"],
            include_raw_data=True,
            include_mcmc_traces=False,  # Skip for simplicity
            include_figures=False,
        )

        raw_data = {"plates": {"plate1": {"wells": {"A1": [1, 2, 3]}}}}
        results = {
            "fitted_params": [{"well_id": "A1", "k_obs": 0.1}],
            "fold_changes": [{"construct": "MUT1", "mean": 2.0}],
            "data_summary": {"n_constructs": 1},
        }

        zip_bytes, manifest = PublicationPackageService.create_publication_package(
            config,
            raw_data,
            results,
        )

        # Verify ZIP is valid
        assert len(zip_bytes) > 0
        with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
            names = zf.namelist()
            assert "manifest.json" in names
            assert "README.md" in names

        # Verify manifest
        assert manifest["package"]["title"] == "Test Analysis"
        assert len(manifest["files"]) > 0

    def test_validate_package_integrity(self):
        """Test package integrity validation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            package_path = Path(tmpdir)
            test_file = package_path / "test.json"
            test_data = b'{"key": "value"}'
            test_file.write_bytes(test_data)

            manifest = {
                "files": [
                    {
                        "filename": "test.json",
                        "sha256": PublicationPackageService.compute_file_hash(test_data),
                        "size_bytes": len(test_data),
                    }
                ]
            }

            is_valid, errors = PublicationPackageService.validate_package_integrity(
                package_path, manifest
            )

            assert is_valid is True
            assert len(errors) == 0

    def test_validate_package_integrity_failure(self):
        """Test package integrity validation with modified file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            package_path = Path(tmpdir)
            test_file = package_path / "test.json"
            test_file.write_bytes(b'{"key": "modified"}')  # Different content

            manifest = {
                "files": [
                    {
                        "filename": "test.json",
                        "sha256": "wronghash",
                        "size_bytes": 100,
                    }
                ]
            }

            is_valid, errors = PublicationPackageService.validate_package_integrity(
                package_path, manifest
            )

            assert is_valid is False
            assert len(errors) > 0


# ========== Validation Service Tests ==========

class TestValidationService:
    """Tests for ValidationService."""

    def test_compute_relative_error(self):
        """Test relative error computation."""
        # Normal case
        assert ValidationService.compute_relative_error(100, 101) == pytest.approx(0.01)

        # Zero expected
        assert ValidationService.compute_relative_error(0, 0) == 0.0
        assert ValidationService.compute_relative_error(0, 5) == 5.0

        # Exact match
        assert ValidationService.compute_relative_error(50, 50) == 0.0

    def test_validate_scalar(self):
        """Test scalar validation."""
        # Valid case
        result = ValidationService.validate_scalar(
            "test_param", 100.0, 100.00001, tolerance=1e-4
        )
        assert result.is_valid is True
        assert result.relative_error < 1e-4

        # Invalid case
        result = ValidationService.validate_scalar(
            "test_param", 100.0, 101.0, tolerance=1e-4
        )
        assert result.is_valid is False
        assert result.relative_error > 1e-4

    def test_validate_scalar_nan(self):
        """Test scalar validation with NaN values."""
        # Both NaN - considered equal
        result = ValidationService.validate_scalar(
            "test", float("nan"), float("nan")
        )
        assert result.is_valid is True

        # One NaN - invalid
        result = ValidationService.validate_scalar(
            "test", float("nan"), 1.0
        )
        assert result.is_valid is False

    def test_validate_array(self):
        """Test array validation."""
        expected = [1.0, 2.0, 3.0]
        actual = [1.00001, 2.00001, 3.00001]

        results = ValidationService.validate_array(
            "test_array", expected, actual, tolerance=1e-4
        )

        assert len(results) == 3
        assert all(r.is_valid for r in results)

    def test_validate_array_length_mismatch(self):
        """Test array validation with length mismatch."""
        expected = [1.0, 2.0, 3.0]
        actual = [1.0, 2.0]

        results = ValidationService.validate_array(
            "test_array", expected, actual
        )

        assert len(results) == 1
        assert results[0].is_valid is False
        assert "length mismatch" in results[0].details

    def test_validate_fitted_parameters(self):
        """Test fitted parameters validation."""
        expected = [
            {"well_id": "A1", "k_obs": 0.1, "f_max": 100.0, "r_squared": 0.99},
            {"well_id": "A2", "k_obs": 0.2, "f_max": 150.0, "r_squared": 0.98},
        ]
        actual = [
            {"well_id": "A1", "k_obs": 0.10001, "f_max": 100.001, "r_squared": 0.99001},
            {"well_id": "A2", "k_obs": 0.20001, "f_max": 150.001, "r_squared": 0.98001},
        ]

        results = ValidationService.validate_fitted_parameters(
            expected, actual, tolerance=1e-3
        )

        # Should have results for each parameter of each well
        assert len(results) > 0
        assert all(r.is_valid for r in results)

    def test_validate_fold_changes(self):
        """Test fold changes validation."""
        expected = [
            {"construct": "MUT1", "mean": 2.0, "ci_lower": 1.5, "ci_upper": 2.5},
        ]
        actual = [
            {"construct": "MUT1", "mean": 2.0001, "ci_lower": 1.5001, "ci_upper": 2.5001},
        ]

        results = ValidationService.validate_fold_changes(
            expected, actual, tolerance=1e-3
        )

        assert len(results) > 0
        assert all(r.is_valid for r in results)

    def test_validate_convergence_diagnostics(self):
        """Test convergence diagnostics validation."""
        expected = {
            "r_hat": {"alpha": 1.001, "beta": 1.002},
            "ess": {"alpha": 4000, "beta": 3900},
        }
        actual = {
            "r_hat": {"alpha": 1.0011, "beta": 1.0021},
            "ess": {"alpha": 4001, "beta": 3901},
        }

        results = ValidationService.validate_convergence_diagnostics(
            expected, actual, tolerance=1e-3
        )

        assert len(results) > 0
        assert all(r.is_valid for r in results)


class TestValidationCertificate:
    """Tests for ValidationCertificate."""

    def test_certificate_to_dict(self):
        """Test certificate serialization."""
        results = [
            ValidationResult("param1", 1.0, 1.0, 0.0, True, 1e-4),
            ValidationResult("param2", 2.0, 2.1, 0.05, False, 1e-4, "Error"),
        ]

        cert = ValidationCertificate(
            package_id="test123",
            validated_at=datetime(2024, 1, 1, 12, 0, 0),
            is_valid=False,
            total_checks=2,
            passed_checks=1,
            failed_checks=1,
            results=results,
            software_version="1.0.0",
            validator_hash="abc123",
        )

        d = cert.to_dict()

        assert d["package_id"] == "test123"
        assert d["is_valid"] is False
        assert d["summary"]["total_checks"] == 2
        assert d["summary"]["pass_rate"] == 0.5
        assert len(d["results"]) == 2


class TestDiffReport:
    """Tests for DiffReport."""

    def test_diff_report_to_markdown(self):
        """Test diff report markdown generation."""
        failed = [
            ValidationResult("param1", 1.0, 1.5, 0.5, False, 1e-4, "Big error"),
        ]

        report = DiffReport(
            generated_at=datetime(2024, 1, 1, 12, 0, 0),
            package_id="test123",
            failed_validations=failed,
            summary="1 failure found",
        )

        md = report.to_markdown()

        assert "# Validation Diff Report" in md
        assert "test123" in md
        assert "param1" in md
        assert "1 failure found" in md


# ========== Export Service Tests (Phase 9.1-9.3) ==========

class TestExportServiceDataExport:
    """Tests for data export methods."""

    def test_export_results_summary_csv(self):
        """Test results summary CSV export."""
        fold_changes = [
            {
                "construct": "MUT1",
                "family": "FAM1",
                "reference": "WT",
                "mean_log2": 1.0,
                "mean_fold": 2.0,
                "ci_lower_log2": 0.5,
                "ci_upper_log2": 1.5,
                "ci_lower_fold": 1.4,
                "ci_upper_fold": 2.8,
                "vif": 1.0,
                "n_replicates": 8,
            },
        ]

        csv_str = ExportService.export_results_summary_csv(fold_changes)

        assert "construct" in csv_str
        assert "MUT1" in csv_str
        assert "mean_log2" in csv_str

    def test_export_results_summary_csv_empty(self):
        """Test empty results export."""
        csv_str = ExportService.export_results_summary_csv([])
        assert csv_str == ""

    def test_export_plate_data_csv_long(self):
        """Test plate data CSV export in long format."""
        plate_data = {
            "wells": {
                "A1": {"construct": "WT", "k_obs": 0.1},
                "A2": {"construct": "MUT1", "k_obs": 0.2},
            },
        }

        csv_str = ExportService.export_plate_data_csv(plate_data, format_type="long")

        assert "well_id" in csv_str
        assert "A1" in csv_str
        assert "A2" in csv_str

    def test_export_plate_data_csv_wide(self):
        """Test plate data CSV export in wide format."""
        plate_data = {
            "time_points": [0, 10, 20],
            "wells": {
                "A1": {"measurements": [0, 50, 75]},
                "A2": {"measurements": [0, 60, 85]},
            },
        }

        csv_str = ExportService.export_plate_data_csv(plate_data, format_type="wide")

        assert "time" in csv_str
        assert "A1" in csv_str
        assert "A2" in csv_str

    def test_export_json_archive(self):
        """Test JSON archive export."""
        data = {
            "raw_data": [{"well": "A1", "values": [1, 2, 3]}],
            "results": {"status": "complete"},
        }

        zip_bytes = ExportService.export_json_archive(data)

        # Verify ZIP is valid
        with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
            names = zf.namelist()
            assert "metadata.json" in names
            assert "raw_data.json" in names
            assert "results.json" in names

            # Verify content
            raw_content = json.loads(zf.read("raw_data.json"))
            assert len(raw_content) == 1
            assert raw_content[0]["well"] == "A1"

    def test_export_json_archive_with_datetime(self):
        """Test JSON archive with datetime objects."""
        data = {
            "event": {"timestamp": datetime(2024, 1, 1, 12, 0, 0)},
        }

        zip_bytes = ExportService.export_json_archive(data)

        with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
            content = json.loads(zf.read("event.json"))
            assert "2024-01-01" in content["timestamp"]


class TestExportServiceExcel:
    """Tests for Excel export methods."""

    def test_export_excel_multisheet(self):
        """Test multi-sheet Excel export."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        sheets = {
            "Data1": [{"col1": "a", "col2": 1}, {"col1": "b", "col2": 2}],
            "Data2": [{"x": 10, "y": 20}],
        }
        metadata = {"project": "Test", "date": "2024-01-01"}

        excel_bytes = ExportService.export_excel_multisheet(sheets, metadata)

        # Verify workbook is valid
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        assert "Metadata" in wb.sheetnames
        assert "Data1" in wb.sheetnames
        assert "Data2" in wb.sheetnames

    def test_export_analysis_excel(self):
        """Test analysis results Excel export."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        raw_data = [{"well": "A1", "time": 0, "value": 100}]
        fitted_params = [{"well_id": "A1", "k_obs": 0.1, "f_max": 100}]
        fold_changes = [{"construct": "MUT1", "mean": 2.0}]

        excel_bytes = ExportService.export_analysis_excel(
            raw_data=raw_data,
            fitted_params=fitted_params,
            fold_changes=fold_changes,
        )

        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        assert "Raw Data" in wb.sheetnames
        assert "Fitted Parameters" in wb.sheetnames
        assert "Fold Changes" in wb.sheetnames


# ========== Integration Tests ==========

class TestPackageCreationAndValidation:
    """Integration tests for package creation and validation."""

    def test_create_and_validate_package(self):
        """Test creating and validating a complete package."""
        # Create package
        config = PublicationPackageConfig(
            title="Integration Test",
            authors=["Test Author"],
            include_mcmc_traces=False,
            include_figures=False,
        )

        raw_data = {
            "plates": {
                "plate1": {"wells": {"A1": [1, 2, 3], "A2": [4, 5, 6]}}
            }
        }
        results = {
            "fitted_params": [
                {"well_id": "A1", "k_obs": 0.1, "f_max": 100, "r_squared": 0.99},
                {"well_id": "A2", "k_obs": 0.2, "f_max": 150, "r_squared": 0.98},
            ],
            "fold_changes": [
                {"construct": "MUT1", "mean": 2.0, "ci_lower": 1.5, "ci_upper": 2.5},
            ],
            "data_summary": {"n_constructs": 1, "n_plates": 1},
        }

        zip_bytes, manifest = PublicationPackageService.create_publication_package(
            config, raw_data, results
        )

        # Extract and validate
        with tempfile.TemporaryDirectory() as tmpdir:
            package_path = Path(tmpdir)
            with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
                zf.extractall(package_path)

            # Verify structure
            assert (package_path / "manifest.json").exists()
            assert (package_path / "README.md").exists()
            assert (package_path / "processed_data" / "fitted_parameters.csv").exists()

            # Validate integrity
            is_valid, errors = PublicationPackageService.validate_package_integrity(
                package_path, manifest
            )
            assert is_valid is True
